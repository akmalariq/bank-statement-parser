"""
Fund Lineage Audit Script
Analyzes transactions across SPAN, BNI, and CASA accounts to trace fund movements.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def load_all_statements(excel_dir: str = 'excel') -> dict:
    """Load all statement Excel files."""
    excel_path = Path(excel_dir)
    
    data = {}
    
    files = {
        'CASA': 'CASA_Combined_Statements.xlsx',
        'BNI': 'BNI_Combined_Statements.xlsx',
        'SPAN': 'SPAN_Combined_Statements.xlsx'
    }
    
    for name, filename in files.items():
        filepath = excel_path / filename
        if filepath.exists():
            data[name] = pd.read_excel(filepath, sheet_name='All_Transactions')
            print(f"✓ Loaded {name}: {len(data[name])} transactions")
        else:
            print(f"⚠ {filename} not found")
            data[name] = pd.DataFrame()
    
    return data


def parse_amount(val) -> float:
    """Parse amount to float."""
    if pd.isna(val):
        return 0.0
    return abs(float(str(val).replace(',', '').replace('-', '').strip() or 0))


def trace_span_withdrawals(data: dict) -> pd.DataFrame:
    """Find all SPAN cash withdrawals (TARIK TUNAI)."""
    span = data.get('SPAN', pd.DataFrame())
    if span.empty:
        return pd.DataFrame()
    
    withdrawals = span[span['Tipe_Transaksi'].str.contains('TARIK', na=False, case=False)].copy()
    withdrawals['Amount'] = withdrawals['Debit'].apply(parse_amount)
    withdrawals['Flow'] = '1. SPAN → Cash'
    
    return withdrawals[['Tanggal', 'Tipe_Transaksi', 'Amount', 'Flow', 'Keterangan']].copy()


def trace_bni_cash_deposits(data: dict) -> pd.DataFrame:
    """Find all BNI cash deposits (SETOR TUNAI)."""
    bni = data.get('BNI', pd.DataFrame())
    if bni.empty:
        return pd.DataFrame()
    
    deposits = bni[bni['Tipe_Transaksi'].str.contains('Setor Tunai', na=False, case=False)].copy()
    deposits['Amount'] = deposits['Kredit'].apply(parse_amount)
    deposits['Flow'] = '2. Cash → BNI'
    
    return deposits[['Tanggal', 'Tipe_Transaksi', 'Amount', 'Flow', 'Keterangan']].copy()


def trace_bni_to_casa_transfers(data: dict) -> pd.DataFrame:
    """Find all BNI to CASA (CIMB) transfers."""
    bni = data.get('BNI', pd.DataFrame())
    if bni.empty:
        return pd.DataFrame()
    
    transfers = bni[
        (bni['Tipe_Transaksi'].str.contains('Transfer', na=False, case=False)) &
        (bni['Penerima'].str.contains('CIMB|INDAH ROSALIA', na=False, case=False))
    ].copy()
    transfers['Amount'] = transfers['Debit'].apply(parse_amount)
    transfers['Flow'] = '3. BNI → CASA'
    
    return transfers[['Tanggal', 'Tipe_Transaksi', 'Amount', 'Flow', 'Penerima']].copy()


def find_suspicious_patterns(data: dict) -> list:
    """Identify suspicious transaction patterns."""
    findings = []
    
    bni = data.get('BNI', pd.DataFrame())
    casa = data.get('CASA', pd.DataFrame())
    
    # 1. Large cash deposits (> 50M)
    if not bni.empty:
        bni['Kredit_Val'] = bni['Kredit'].apply(parse_amount)
        large_deposits = bni[
            (bni['Tipe_Transaksi'].str.contains('Setor Tunai', na=False, case=False)) &
            (bni['Kredit_Val'] > 50000000)
        ]
        for _, row in large_deposits.iterrows():
            findings.append({
                'Type': 'Large Cash Deposit',
                'Account': 'BNI',
                'Date': row['Tanggal'],
                'Amount': row['Kredit_Val'],
                'Description': row.get('Keterangan', ''),
                'Risk': 'HIGH',
                'Note': 'Cash deposit > Rp 50M requires documentation'
            })
    
    # 2. Round number transactions
    if not bni.empty:
        bni['Debit_Val'] = bni['Debit'].apply(parse_amount)
        round_txns = bni[
            ((bni['Debit_Val'] > 0) & (bni['Debit_Val'] % 10000000 == 0)) |
            ((bni['Kredit_Val'] > 0) & (bni['Kredit_Val'] % 10000000 == 0))
        ]
        for _, row in round_txns.head(20).iterrows():  # Limit to 20
            amt = row['Debit_Val'] if row['Debit_Val'] > 0 else row['Kredit_Val']
            if amt >= 50000000:
                findings.append({
                    'Type': 'Round Number',
                    'Account': 'BNI',
                    'Date': row['Tanggal'],
                    'Amount': amt,
                    'Description': row.get('Tipe_Transaksi', ''),
                    'Risk': 'MEDIUM',
                    'Note': 'Round numbers may indicate structured transactions'
                })
    
    # 3. Same-day large movements
    if not bni.empty:
        bni_grouped = bni.groupby('Tanggal').agg({
            'Debit_Val': 'sum',
            'Kredit_Val': 'sum'
        }).reset_index()
        high_volume_days = bni_grouped[
            (bni_grouped['Debit_Val'] > 100000000) | 
            (bni_grouped['Kredit_Val'] > 100000000)
        ]
        for _, row in high_volume_days.iterrows():
            findings.append({
                'Type': 'High Volume Day',
                'Account': 'BNI',
                'Date': row['Tanggal'],
                'Amount': max(row['Debit_Val'], row['Kredit_Val']),
                'Description': f"Debit: {row['Debit_Val']:,.0f} | Credit: {row['Kredit_Val']:,.0f}",
                'Risk': 'MEDIUM',
                'Note': 'High transaction volume on single day'
            })
    
    return findings


def calculate_lineage_summary(data: dict) -> dict:
    """Calculate summary statistics for fund lineage."""
    span_withdraw = trace_span_withdrawals(data)
    bni_deposit = trace_bni_cash_deposits(data)
    bni_to_casa = trace_bni_to_casa_transfers(data)
    
    return {
        'span_withdrawals': {
            'count': len(span_withdraw),
            'total': span_withdraw['Amount'].sum() if not span_withdraw.empty else 0
        },
        'bni_cash_deposits': {
            'count': len(bni_deposit),
            'total': bni_deposit['Amount'].sum() if not bni_deposit.empty else 0
        },
        'bni_to_casa': {
            'count': len(bni_to_casa),
            'total': bni_to_casa['Amount'].sum() if not bni_to_casa.empty else 0
        }
    }


def generate_audit_report(data: dict, output_file: str = 'excel/Audit_Report.xlsx'):
    """Generate comprehensive audit report Excel."""
    
    # Collect all lineage data
    span_withdraw = trace_span_withdrawals(data)
    bni_deposit = trace_bni_cash_deposits(data)
    bni_to_casa = trace_bni_to_casa_transfers(data)
    suspicious = find_suspicious_patterns(data)
    summary = calculate_lineage_summary(data)
    
    # Combine all lineage transactions
    all_lineage = []
    
    for _, row in span_withdraw.iterrows():
        all_lineage.append({
            'Flow': row['Flow'],
            'Date': row['Tanggal'],
            'Type': row['Tipe_Transaksi'],
            'Amount': row['Amount'],
            'Details': row.get('Keterangan', '')
        })
    
    for _, row in bni_deposit.iterrows():
        all_lineage.append({
            'Flow': row['Flow'],
            'Date': row['Tanggal'],
            'Type': row['Tipe_Transaksi'],
            'Amount': row['Amount'],
            'Details': row.get('Keterangan', '')
        })
    
    for _, row in bni_to_casa.iterrows():
        all_lineage.append({
            'Flow': row['Flow'],
            'Date': row['Tanggal'],
            'Type': row['Tipe_Transaksi'],
            'Amount': row['Amount'],
            'Details': row.get('Penerima', '')
        })
    
    lineage_df = pd.DataFrame(all_lineage)
    suspicious_df = pd.DataFrame(suspicious)
    
    # Create Excel
    wb = Workbook()
    hdr_fill = PatternFill(start_color='2F5496', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF')
    flow_fills = {
        '1. SPAN → Cash': PatternFill(start_color='FFF2CC', fill_type='solid'),
        '2. Cash → BNI': PatternFill(start_color='DEEBF7', fill_type='solid'),
        '3. BNI → CASA': PatternFill(start_color='E2EFDA', fill_type='solid')
    }
    risk_fills = {
        'HIGH': PatternFill(start_color='FF6B6B', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFE699', fill_type='solid'),
        'LOW': PatternFill(start_color='C6EFCE', fill_type='solid')
    }
    
    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    
    summary_rows = [
        ['FUND LINEAGE AUDIT REPORT', '', ''],
        [f'Generated: {datetime.now():%Y-%m-%d %H:%M}', '', ''],
        ['', '', ''],
        ['FUND FLOW SUMMARY', 'Count', 'Total Amount'],
        ['1. SPAN Withdrawals (TARIK TUNAI)', summary['span_withdrawals']['count'], 
         summary['span_withdrawals']['total']],
        ['2. BNI Cash Deposits (SETOR TUNAI)', summary['bni_cash_deposits']['count'], 
         summary['bni_cash_deposits']['total']],
        ['3. BNI → CASA Transfers', summary['bni_to_casa']['count'], 
         summary['bni_to_casa']['total']],
        ['', '', ''],
        ['FLOW DIAGRAM:', '', ''],
        ['  SPAN (Govt)', '  ───►', '  BNI (Personal)  ───►  CASA (CIMB)'],
        ['  TARIK TUNAI', '', '  SETOR TUNAI           BI-FAST'],
        ['', '', ''],
        ['SUSPICIOUS FINDINGS', len(suspicious), 'transactions flagged'],
    ]
    
    for i, row in enumerate(summary_rows, 1):
        for j, val in enumerate(row, 1):
            c = ws_sum.cell(row=i, column=j, value=val)
            if i == 1:
                c.font = Font(size=16, bold=True, color='2F5496')
            elif i == 4:
                c.fill, c.font = hdr_fill, hdr_font
            if j == 3 and i >= 5 and i <= 7:
                c.number_format = '#,##0'
    
    ws_sum.column_dimensions['A'].width = 40
    ws_sum.column_dimensions['B'].width = 15
    ws_sum.column_dimensions['C'].width = 20
    
    # Sheet 2: All Lineage
    ws_lin = wb.create_sheet('Fund_Lineage')
    if not lineage_df.empty:
        hdrs = ['Flow', 'Date', 'Type', 'Amount', 'Details']
        for col, h in enumerate(hdrs, 1):
            c = ws_lin.cell(row=1, column=col, value=h)
            c.fill, c.font = hdr_fill, hdr_font
        
        for i, row in lineage_df.iterrows():
            for col, k in enumerate(hdrs, 1):
                c = ws_lin.cell(row=i+2, column=col, value=row.get(k, ''))
                c.fill = flow_fills.get(row['Flow'], PatternFill())
                if k == 'Amount':
                    c.number_format = '#,##0'
        
        ws_lin.freeze_panes = 'A2'
    
    # Sheet 3: Suspicious
    ws_susp = wb.create_sheet('Suspicious')
    if not suspicious_df.empty:
        hdrs = ['Type', 'Account', 'Date', 'Amount', 'Description', 'Risk', 'Note']
        for col, h in enumerate(hdrs, 1):
            c = ws_susp.cell(row=1, column=col, value=h)
            c.fill, c.font = hdr_fill, hdr_font
        
        for i, (_, row) in enumerate(suspicious_df.iterrows(), 2):
            for col, k in enumerate(hdrs, 1):
                c = ws_susp.cell(row=i, column=col, value=row.get(k, ''))
                if k == 'Risk':
                    c.fill = risk_fills.get(row.get('Risk', ''), PatternFill())
                if k == 'Amount':
                    c.number_format = '#,##0'
        
        ws_susp.freeze_panes = 'A2'
    
    # Auto-adjust columns
    for sheet in [ws_sum, ws_lin, ws_susp]:
        for column in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    wb.save(output_file)
    return output_file


def run_audit():
    """Main audit function."""
    print("=" * 70)
    print("FUND LINEAGE AUDIT")
    print("=" * 70)
    print()
    
    # Load data
    print("📂 Loading statements...")
    data = load_all_statements()
    print()
    
    # Calculate summary
    print("📊 Analyzing fund flow...")
    summary = calculate_lineage_summary(data)
    
    print()
    print("=" * 70)
    print("FUND FLOW SUMMARY")
    print("=" * 70)
    print(f"1. SPAN Withdrawals:     {summary['span_withdrawals']['count']:>3} txns | Rp {summary['span_withdrawals']['total']:>15,.0f}")
    print(f"2. BNI Cash Deposits:    {summary['bni_cash_deposits']['count']:>3} txns | Rp {summary['bni_cash_deposits']['total']:>15,.0f}")
    print(f"3. BNI → CASA Transfers: {summary['bni_to_casa']['count']:>3} txns | Rp {summary['bni_to_casa']['total']:>15,.0f}")
    print()
    
    # Find suspicious patterns
    print("🔍 Scanning for suspicious patterns...")
    suspicious = find_suspicious_patterns(data)
    print(f"   Found {len(suspicious)} suspicious findings")
    
    if suspicious:
        print()
        high_risk = [s for s in suspicious if s['Risk'] == 'HIGH']
        if high_risk:
            print(f"   ⚠️  HIGH RISK: {len(high_risk)} findings")
    
    print()
    
    # Generate report
    print("📝 Generating audit report...")
    output = generate_audit_report(data)
    print(f"   ✅ Saved to: {output}")
    print()
    print("=" * 70)


if __name__ == "__main__":
    run_audit()
